import openpyxl

# Create workbook and sheets
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Login'  # Changed from 'Sheet1' to 'Login'

# Login Sheet - Header
ws['A1'] = 'Email'
ws['B1'] = 'Password'
ws['C1'] = 'Expected Result'

# Login Sheet - Test Data (Invalid test first, then valid test)
ws['A2'] = 'm46328690@gmail.com'
ws['B2'] = 'wrongpassword'
ws['C2'] = 'Fail'

ws['A3'] = 'm46328690@gmail.com'
ws['B3'] = '1234567'
ws['C3'] = 'Pass'

# SignUp Sheet - Create and setup
ws2 = wb.create_sheet('SignUp')
ws2['A1'] = 'First Name'
ws2['B1'] = 'Last Name'
ws2['C1'] = 'Email'
ws2['D1'] = 'Password'
ws2['E1'] = 'Expected Result'

# SignUp - Single test case
ws2['A2'] = 'Muhammad'
ws2['B2'] = 'Ilyas'
ws2['C2'] = 'newuser54321@test.com'
ws2['D2'] = '1234567'
ws2['E2'] = 'Pass'

# AddContact Sheet - Create and setup
ws3 = wb.create_sheet('AddContact')
ws3['A1'] = 'First Name'
ws3['B1'] = 'Last Name'
ws3['C1'] = 'Birthdate'
ws3['D1'] = 'Email'
ws3['E1'] = 'Phone'
ws3['F1'] = 'Street1'
ws3['G1'] = 'Street2'
ws3['H1'] = 'City'
ws3['I1'] = 'State/Province'
ws3['J1'] = 'Postal Code'
ws3['K1'] = 'Country'
ws3['L1'] = 'Expected Result'

# AddContact - Single test case
ws3['A2'] = 'Faris'
ws3['B2'] = 'Hamza'
ws3['C2'] = '2003-04-25'
ws3['D2'] = 'faris@gmail.com'
ws3['E2'] = '03114593627'
ws3['F2'] = 'College Road'
ws3['G2'] = 'Near Lajna Chok'
ws3['H2'] = 'Lahore'
ws3['I2'] = 'Punjab'
ws3['J2'] = '34325'
ws3['K2'] = 'Pakistan'
ws3['L2'] = 'Pass'

# UpdateContact Sheet - Create and setup (Only 1 test case)
ws4 = wb.create_sheet('UpdateContact')
ws4['A1'] = 'First Name'
ws4['B1'] = 'Last Name'
ws4['C1'] = 'Birthdate'
ws4['D1'] = 'Email'
ws4['E1'] = 'Phone'
ws4['F1'] = 'Street1'
ws4['G1'] = 'Street2'
ws4['H1'] = 'City'
ws4['I1'] = 'State/Province'
ws4['J1'] = 'Postal Code'
ws4['K1'] = 'Country'
ws4['L1'] = 'Expected Result'

# UpdateContact - Single test case
ws4['A2'] = 'Faris Updated'
ws4['B2'] = 'Hamza Updated'
ws4['C2'] = '2003-05-15'
ws4['D2'] = 'faris.updated@gmail.com'
ws4['E2'] = '03229876543'
ws4['F2'] = 'Updated Street'
ws4['G2'] = 'Near Updated Location'
ws4['H2'] = 'Islamabad'
ws4['I2'] = 'Islamabad Capital'
ws4['J2'] = '44000'
ws4['K2'] = 'Pakistan'
ws4['L2'] = 'Pass'

# DeleteContact Sheet - Create and setup (Only 1 test case)
ws5 = wb.create_sheet('DeleteContact')
ws5['A1'] = 'First Name'
ws5['B1'] = 'Last Name'
ws5['C1'] = 'Expected Result'

# DeleteContact - Single test case
ws5['A2'] = 'Faris Updated'
ws5['B2'] = 'Hamza Updated'
ws5['C2'] = 'Pass'

# Save to both locations
wb.save('src/test/resources/testdata/Data.xlsx')
wb.save('src/main/resources/testdata/Data.xlsx')

print('Excel file created successfully!')
print(f'Login Row 2 (Invalid): Expected Result = {ws["C2"].value}')
print(f'Login Row 3 (Valid): Expected Result = {ws["C3"].value}')
print(f'SignUp Row 2 (Single test): Expected Result = {ws2["E2"].value}')
print(f'AddContact Row 2 (Single test): Expected Result = {ws3["L2"].value}')
print(f'UpdateContact Row 2: Expected Result = {ws4["L2"].value} - Only 1 test case')
print(f'DeleteContact Row 2: Expected Result = {ws5["C2"].value} - Only 1 test case')
